"""Install the built wheel offline and verify import plus CLI help."""

from __future__ import annotations

import hashlib
import os
import socket
import subprocess
import sys
import tomllib
from importlib.metadata import version
from pathlib import Path

import pytest
from packaging.utils import canonicalize_name, parse_wheel_filename


def _network_denied(*_args: object, **_kwargs: object) -> None:
    raise RuntimeError("network access denied during package test")


def _run(command: list[str], *, env: dict[str, str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        command,
        check=True,
        capture_output=True,
        text=True,
        env=env,
        timeout=60,
    )


def test_clean_wheel_install_offline(
    repository_root: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(socket, "create_connection", _network_denied)
    monkeypatch.setattr(socket.socket, "connect", _network_denied)
    monkeypatch.setattr(socket.socket, "connect_ex", _network_denied)
    environment = {
        name: os.environ[name]
        for name in (
            "COMSPEC",
            "HOME",
            "LANG",
            "PATH",
            "PATHEXT",
            "SYSTEMROOT",
            "TEMP",
            "TMP",
            "TMPDIR",
            "WINDIR",
        )
        if name in os.environ
    }
    environment["HOME"] = str(tmp_path / "empty-home")
    environment["UV_NO_CONFIG"] = "1"
    environment["UV_OFFLINE"] = "1"
    environment["UV_CACHE_DIR"] = str(tmp_path / "empty-uv-cache")
    wheels = sorted(
        (repository_root / "dist").glob(f"survey_scribe-{version('survey-scribe')}-*.whl")
    )
    assert len(wheels) == 1, "Build exactly one current-version wheel before package tests."
    wheel = wheels[0]
    virtual_environment = tmp_path / "venv"
    wheelhouse = repository_root / ".cache/wheelhouse"
    assert wheelhouse.is_dir(), "Prepare the locked test wheelhouse before package tests."
    locked = tomllib.loads((repository_root / "uv.lock").read_text(encoding="utf-8"))
    locked_packages = {
        (canonicalize_name(package["name"]), package["version"]): package
        for package in locked["package"]
    }
    dependency_wheels = sorted(wheelhouse.glob("*.whl"))
    assert dependency_wheels
    hashed_requirements = [
        f"{wheel.as_uri()} --hash=sha256:{hashlib.sha256(wheel.read_bytes()).hexdigest()}"
    ]
    for dependency in dependency_wheels:
        name, dependency_version, _build, _tags = parse_wheel_filename(dependency.name)
        digest = f"sha256:{hashlib.sha256(dependency.read_bytes()).hexdigest()}"
        locked_package = locked_packages[(canonicalize_name(name), str(dependency_version))]
        assert digest in {item["hash"] for item in locked_package["wheels"]}
        hashed_requirements.append(f"{name}=={dependency_version} --hash={digest}")
    requirements = tmp_path / "locked-wheelhouse.txt"
    requirements.write_text("\n".join(hashed_requirements) + "\n", encoding="utf-8")

    _run(
        ["uv", "venv", "--python", sys.executable, str(virtual_environment)],
        env=environment,
    )
    python = virtual_environment / ("Scripts/python.exe" if os.name == "nt" else "bin/python")
    _run(
        [
            "uv",
            "pip",
            "install",
            "--python",
            str(python),
            "--no-index",
            "--find-links",
            str(wheelhouse),
            "--require-hashes",
            "--requirement",
            str(requirements),
        ],
        env=environment,
    )
    guarded_import = """
import asyncio
import socket

def deny_network(*args, **kwargs):
    raise RuntimeError("network access denied during package smoke test")

socket.create_connection = deny_network
socket.socket.connect = deny_network
import survey_scribe
import tempfile
from datetime import date
from pathlib import Path
from importlib.metadata import version
from openpyxl import Workbook
from pydantic import BaseModel, ConfigDict
from survey_scribe import QuestionnaireRouter, RoutedSurveySVIS
from survey_scribe.cli import main
from survey_scribe.config import GenerationConfig, RetryConfig, SurveyScribeConfig
from survey_scribe.models import DataType, SurveySVIS, SurveyVariable
from survey_scribe.providers.azure import AzureOpenAIProvider
from survey_scribe.providers.base import ConcurrencyLimiter, ProviderMessage
from survey_scribe.providers.capabilities import CapabilityEvidence, ModelCapabilities
from survey_scribe.results import ExtractionResult, ResultStatus
from survey_scribe.serialization.artifacts import write_result
from survey_scribe.serialization import legacy_json_bytes, legacy_payload
from survey_scribe.serialization.routing import ArtifactManifestV2, parse_artifact_manifest
from survey_scribe.sources import SourceLimits
from survey_scribe.sources.chunking import ConservativeTokenEstimator
from survey_scribe.sources.ocr import APPROVED_OCR_ARTIFACTS
from survey_scribe.sources.registry import SourceRegistry
assert survey_scribe.__version__ == version("survey-scribe")
assert SurveyScribeConfig().provider == "openai"
assert ExtractionResult is not None
assert write_result is not None
assert legacy_payload({"safe": True}) == {"safe": True}
assert SourceLimits().max_pages == 2000
assert ConservativeTokenEstimator().estimate("abc") == 3
assert len(APPROVED_OCR_ARTIFACTS) == 2

class WheelAnswer(BaseModel):
    model_config = ConfigDict(extra="forbid")
    value: int

captured_headers = []
def completion(**kwargs):
    captured_headers.append(dict(kwargs["extra_headers"]))
    return WheelAnswer(value=7)

capabilities = ModelCapabilities(
    provider="azure_openai",
    model="wheel-deployment",
    structured_output=True,
    strict_schema=True,
    max_input_tokens=4096,
    max_output_tokens=1024,
    supported_generation_settings=frozenset({"temperature", "max_output_tokens", "seed"}),
    evidence=CapabilityEvidence.configuration_only,
    tested_sdk_version="wheel-smoke",
)
azure_provider = AzureOpenAIProvider(
    deployment=capabilities.model,
    azure_endpoint="https://gateway.example/azure",
    api_version="api-version",
    metadata_headers={"X-Synthetic-Route": "wheel-smoke"},
    sensitive_headers_callback=lambda: {"X-Synthetic-Aux-Key": "aux-value"},
    required_headers=("X-Synthetic-Route", "X-Synthetic-Aux-Key"),
    capabilities=capabilities,
    completion=completion,
)
azure_response = asyncio.run(
    azure_provider.generate(
        messages=(ProviderMessage(role="user", content="synthetic"),),
        response_model=WheelAnswer,
        generation=GenerationConfig(max_output_tokens=1024),
        retry=RetryConfig(max_attempts=1),
        limiter=ConcurrencyLimiter(1),
    )
)
assert azure_response.output.value == 7
assert captured_headers == [{
    "X-Synthetic-Route": "wheel-smoke",
    "X-Synthetic-Aux-Key": "aux-value",
}]
asyncio.run(azure_provider.aclose())

with tempfile.TemporaryDirectory() as temporary_directory:
    output_directory = Path(temporary_directory)
    source = output_directory / "questionnaire.xlsx"
    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    survey.append(["type", "name", "label", "relevant"])
    survey.append(["select_one yes_no", "consent", "Consent?", ""])
    survey.append(["integer", "age", "Age", "${consent} = 'yes'"])
    choices = workbook.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    choices.append(["yes_no", "yes", "Yes"])
    choices.append(["yes_no", "no", "No"])
    settings = workbook.create_sheet("settings")
    settings.append(["form_title", "form_id", "country_code", "year"])
    settings.append(["Package Smoke", "package_smoke", "TST", 2026])
    workbook.save(source)
    svis = SurveySVIS(
        survey_id="TST_2026_WHEEL",
        country_code="TST",
        year=2026,
        survey_name="Wheel smoke",
        variables=[
            SurveyVariable(raw_name="consent", data_type=DataType.categorical_single, extraction_confidence=1.0),
            SurveyVariable(raw_name="age", data_type=DataType.numeric, extraction_confidence=1.0),
        ],
        source_file=source.name,
        source_format="xlsx",
        extraction_date=date(2026, 9, 1),
    )
    registry = SourceRegistry.default()
    binding = registry.convert_with_native(source, svis).source_binding
    routed = QuestionnaireRouter(None, sources=registry).route(
        source,
        svis,
        source_binding=binding,
    )
    assert routed.status is ResultStatus.success
    assert isinstance(routed.output, RoutedSurveySVIS)
    assert all(variable.routing_node_id is not None for variable in routed.output.variables)
    consent, age = routed.output.variables
    nodes = {node.node_id: node for node in routed.output.routing_graph.nodes}
    age_activation = nodes[age.routing_node_id].activation_condition
    assert age_activation is not None
    assert age_activation.operator.value == "equals"
    assert age_activation.question_node_id == consent.routing_node_id
    assert age_activation.value == "yes"
    assert any(
        record.observation.origin.value == "native_parser"
        for record in routed.output.routing_graph.routing_audit.evidence
    )
    inventory_node_ids = {
        item.node_id for item in routed.output.routing_graph.routing_audit.inventory
    }
    assert all(variable.routing_node_id in inventory_node_ids for variable in routed.output.variables)
    written = routed.write(output_directory)
    manifest_path = next(item.path for item in written.artifacts if item.kind == "manifest")
    assert isinstance(parse_artifact_manifest(manifest_path.read_bytes()), ArtifactManifestV2)
    legacy_path = output_directory / "TST_2026_WHEEL_svis.json"
    assert legacy_path.read_bytes() == legacy_json_bytes(svis)
    SurveySVIS.model_validate_json(legacy_path.read_bytes())
try:
    main(["--help"])
except SystemExit as exc:
    assert exc.code == 0
"""
    imported = _run(
        [str(python), "-I", "-c", guarded_import],
        env=environment,
    )
    executable = virtual_environment / (
        "Scripts/survey-scribe.exe" if os.name == "nt" else "bin/survey-scribe"
    )
    network_guard = tmp_path / "network-guard"
    network_guard.mkdir()
    network_guard.joinpath("sitecustomize.py").write_text(
        "import socket\n"
        "def deny_network(*args, **kwargs):\n"
        "    raise RuntimeError('network access denied during package CLI smoke test')\n"
        "socket.create_connection = deny_network\n"
        "socket.socket.connect = deny_network\n",
        encoding="utf-8",
    )
    runtime_environment = environment | {
        "PYTHONNOUSERSITE": "1",
        "PYTHONPATH": str(network_guard),
    }
    help_result = _run(
        [str(executable), "--help"],
        env=runtime_environment,
    )
    schema_result = _run(
        [str(executable), "schema", "export", "routing"],
        env=runtime_environment,
    )
    assert imported.returncode == 0
    assert "survey-scribe" in help_result.stdout
    expected_schema = (
        repository_root / "tests/fixtures/routing/schema/questionnaire-routing-graph-v1.0.json"
    ).read_text(encoding="utf-8")
    assert schema_result.stdout == expected_schema
    assert schema_result.stderr == ""

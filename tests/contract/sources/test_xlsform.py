"""Bounded XLSForm parsing, preservation, and native-expression contracts."""

from __future__ import annotations

import json
import time
import zipfile
from dataclasses import replace
from datetime import date
from pathlib import Path

import pytest

from survey_scribe import SurveyScribe
from survey_scribe.models.svis import DataType, SurveySVIS, SurveyVariable
from survey_scribe.providers.capabilities import CapabilityEvidence, ModelCapabilities
from survey_scribe.providers.testing import DeterministicFakeProvider
from survey_scribe.results import DiagnosticCode, ResultStatus
from survey_scribe.routing.contracts import ConditionOperator, NodeKind
from survey_scribe.routing.native import NativeRoutingItem, prepare_native_routing
from survey_scribe.sources.base import (
    DEFAULT_SOURCE_LIMITS,
    SourceBundle,
    SourceFormatError,
    SourceLimitError,
    SourceProvenance,
    SourceSecurityError,
    SourceTable,
    SourceTimeoutError,
)
from survey_scribe.sources.registry import SourceRegistry
from survey_scribe.sources.xlsform import (
    XLSFORM_SUPPORT_MATRIX,
    XLSFORM_SUPPORT_MATRIX_VERSION,
    XlsFormAdapter,
    _choice_reference,
    _parse_expression,
    _project_expression,
    _scalar,
    _split_top_level,
    _strip_outer_parentheses,
    _supported_question_type,
    _table_records,
)

_FIXTURES = Path(__file__).parents[2] / "fixtures" / "sources" / "xlsform"


def _workbook(path: Path, survey_rows: list[list[object]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    survey = workbook.active
    assert survey is not None
    survey.title = "survey"
    for row in survey_rows:
        survey.append(row)
    choices = workbook.create_sheet("choices")
    choices.append(["list_name", "name", "label", "label::French (fr)"])
    choices.append(["yes_no", "yes", "Yes", "Oui"])
    choices.append(["yes_no", "no", "No", "Non"])
    settings = workbook.create_sheet("settings")
    settings.append(["form_title", "form_id", "version"])
    settings.append(["Synthetic roster", "synthetic_roster", "20260901"])
    workbook.save(path)


def _svis(path: Path, *raw_names: str) -> SurveySVIS:
    return SurveySVIS(
        survey_id="TST_2026_XLSFORM",
        country_code="TST",
        year=2026,
        survey_name="Synthetic XLSForm",
        variables=[
            SurveyVariable(
                raw_name=name,
                data_type=DataType.numeric,
                extraction_confidence=1.0,
            )
            for name in raw_names
        ],
        source_file=path.name,
        source_format="xlsx",
        extraction_date=date(2026, 9, 1),
    )


def _survey_only(path: Path, rows: list[list[object]], *, title: str = "survey") -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    survey = workbook.active
    assert survey is not None
    survey.title = title
    for row in rows:
        survey.append(row)
    workbook.save(path)


def _versioned_fixture(tmp_path: Path) -> tuple[Path, Path]:
    from openpyxl import Workbook

    fixture = json.loads((_FIXTURES / "core-v1.json").read_text(encoding="utf-8"))
    workbook = Workbook()
    initial = workbook.active
    assert initial is not None
    workbook.remove(initial)
    for name, rows in fixture["sheets"].items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    path = tmp_path / "core-v1.xlsx"
    workbook.save(path)
    companion = tmp_path / "places-v1.csv"
    companion.write_bytes((_FIXTURES / companion.name).read_bytes())
    return path, companion


def _capabilities() -> ModelCapabilities:
    return ModelCapabilities(
        provider="fake",
        model="xlsform-native-v1",
        structured_output=True,
        strict_schema=True,
        max_input_tokens=200_000,
        max_output_tokens=8_192,
        supported_generation_settings=frozenset({"temperature", "max_output_tokens", "seed"}),
        evidence=CapabilityEvidence.verified,
        tested_sdk_version="fake-1",
    )


def test_support_matrix_is_versioned_and_names_exact_projection_boundaries() -> None:
    assert XLSFORM_SUPPORT_MATRIX_VERSION == "1.0"
    assert XLSFORM_SUPPORT_MATRIX["reference_comparisons"] == "exact"
    assert XLSFORM_SUPPORT_MATRIX["selected"] == "exact"
    assert XLSFORM_SUPPORT_MATRIX["boolean_and_or_not"] == "exact"
    assert XLSFORM_SUPPORT_MATRIX["functions_other_than_selected"] == "opaque"
    assert XLSFORM_SUPPORT_MATRIX["arithmetic"] == "opaque"
    assert XLSFORM_SUPPORT_MATRIX["constraints"] == "preserved_not_flow"
    assert XLSFORM_SUPPORT_MATRIX["calculations"] == "preserved_not_flow"
    assert XLSFORM_SUPPORT_MATRIX["choice_filters"] == "preserved_not_flow"
    assert XLSFORM_SUPPORT_MATRIX["tested_profile"] == (
        "XLSForm 1.3 core; ODK-compatible workbook structure"
    )
    assert "select_one_from_file" in XLSFORM_SUPPORT_MATRIX["question_types"]
    assert "XLSFORM_REFERENCE_UNRESOLVED" in XLSFORM_SUPPORT_MATRIX["diagnostic_codes"]
    assert DiagnosticCode.xlsform_type_unsupported == "XLSFORM_TYPE_UNSUPPORTED"


@pytest.mark.parametrize(
    "question_type",
    XLSFORM_SUPPORT_MATRIX["question_types"].split(","),
)
def test_support_matrix_question_types_match_parser(question_type: str) -> None:
    assert _supported_question_type(question_type)


def test_real_xlsform_preserves_sheets_groups_repeat_and_typed_relevance(
    tmp_path: Path,
) -> None:
    path = tmp_path / "roster.xlsx"
    _workbook(
        path,
        [
            [
                "type",
                "name",
                "label",
                "label::French (fr)",
                "relevant",
                "constraint",
                "calculation",
                "choice_filter",
            ],
            ["select_one yes_no", "consent", "Consent?", "Consentement ?", "", "", "", ""],
            ["begin group", "roster", "Roster", "Liste", "${consent} = 'yes'", "", "", ""],
            ["begin repeat", "member", "Member", "Membre", "", "", "", ""],
            [
                "integer",
                "age",
                "Age",
                "Age",
                "${consent} = 'yes' and ${age} >= 0",
                ". >= 0",
                "",
                "",
            ],
            ["calculate", "adult", "", "", "", "", "if(${age} >= 18, 1, 0)", ""],
            ["end repeat", "", "", "", "", "", "", ""],
            ["end group", "", "", "", "", "", "", ""],
        ],
    )

    svis = _svis(path, "consent", "age", "adult")
    converted = SourceRegistry.default().convert_with_native(path, svis)

    assert converted.native is not None
    native = converted.native
    assert native.complete is True
    assert [item.kind for item in native.items].count(NodeKind.repeat_group) == 1
    assert [item.raw_reference for item in native.items].count("member") == 1
    age = next(item for item in native.items if item.raw_reference == "age")
    repeat = next(item for item in native.items if item.kind is NodeKind.repeat_group)
    assert age.repeat_group_local_id == repeat.local_id
    assert len(native.activations) == 2
    assert native.activations[0].expression.projection.operator is ConditionOperator.equals
    assert native.activations[1].expression.projection.operator is ConditionOperator.all
    assert [
        (activation.source_span.row_start, activation.source_span.row_end)
        for activation in native.activations
    ] == [(3, 3), (5, 5)]
    assert [
        (transition.source_span.row_start, transition.source_span.row_end)
        for transition in native.transitions
    ] == [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5)]
    assert tuple(
        reference.source_item_id for reference in native.activations[1].expression.references
    ) == (
        "consent",
        "age",
    )

    records = {(record.collection, dict(record.values).get("name")) for record in native.records}
    assert ("survey", "age") in records
    assert ("choices", "yes") in records
    assert ("settings", None) in records
    survey_age = next(
        record
        for record in native.records
        if record.collection == "survey" and dict(record.values).get("name") == "age"
    )
    assert dict(survey_age.values)["constraint"] == ". >= 0"
    calculation = next(
        record
        for record in native.records
        if record.collection == "survey" and dict(record.values).get("name") == "adult"
    )
    assert dict(calculation.values)["calculation"] == "if(${age} >= 18, 1, 0)"
    assert all(item.raw_reference != "adult" for item in native.items)
    age_transition = next(
        transition
        for transition in native.transitions
        if transition.source_local_id == age.local_id
    )
    assert age_transition.target_local_id == "xlsform:terminal"
    local_ids = {item.local_id for item in native.items}
    assert all(
        transition.source_local_id in local_ids and transition.target_local_id in local_ids
        for transition in native.transitions
    )
    prepared = prepare_native_routing(native, converted.document, svis)
    assert prepared.inventory.variable_node_ids[-1] is None
    assert {(span.row_start, span.row_end) for span in prepared.evidence.source_spans} == {
        (1, 1),
        (2, 2),
        (3, 3),
        (4, 4),
        (5, 5),
    }
    assert all(span.source_quote.count("\n") == 0 for span in prepared.evidence.source_spans)


def test_unsupported_function_and_arithmetic_stay_typed_and_opaque(tmp_path: Path) -> None:
    path = tmp_path / "opaque.xlsx"
    _workbook(
        path,
        [
            ["type", "name", "label", "relevant"],
            ["integer", "age", "Age", ""],
            ["text", "note", "Note", "count-selected(${age}) + 1 > 2"],
        ],
    )

    converted = SourceRegistry.default().convert_with_native(path, _svis(path, "age", "note"))

    assert converted.native is not None
    expression = converted.native.activations[0].expression
    assert expression.exact_expression == "count-selected(${age}) + 1 > 2"
    assert expression.projection.operator is ConditionOperator.opaque
    assert tuple(reference.source_item_id for reference in expression.references) == ("age",)


def test_xlsform_rejects_excel_formulas_macros_and_escaped_external_choices(
    tmp_path: Path,
) -> None:
    formula = tmp_path / "formula.xlsx"
    _workbook(
        formula,
        [
            ["type", "name", "label", "calculation"],
            ["calculate", "unsafe", "Unsafe", "=1+1"],
        ],
    )
    with pytest.raises(SourceSecurityError, match="formula"):
        SourceRegistry.default().convert_with_native(formula, _svis(formula, "unsafe"))

    for unsafe_name, member, message in (
        ("macro.xlsx", "xl/vbaProject.bin", "macro"),
        ("external-link.xlsx", "xl/externalLinks/externalLink1.xml", "external link"),
    ):
        unsafe = tmp_path / unsafe_name
        _workbook(unsafe, [["type", "name"], ["text", "safe"]])
        with zipfile.ZipFile(unsafe, "a") as archive:
            archive.writestr(member, b"unsafe")
        with pytest.raises(SourceSecurityError, match=message):
            SourceRegistry.default().convert_with_native(unsafe, _svis(unsafe, "safe"))

    escaped = tmp_path / "escaped.xlsx"
    _workbook(
        escaped,
        [
            ["type", "name", "label"],
            ["select_one_from_file ../outside.csv", "place", "Place"],
        ],
    )
    with pytest.raises(SourceSecurityError, match="external-choice"):
        SourceRegistry.default().convert_with_native(escaped, _svis(escaped, "place"))

    companion = tmp_path / "places.csv"
    companion.write_text("list_name,name,label\nplaces,a,Place A\n", encoding="utf-8")
    safe = tmp_path / "safe.xlsx"
    _workbook(
        safe,
        [
            ["type", "name", "label"],
            ["select_one_from_file places.csv", "place", "Place"],
        ],
    )
    converted = SourceRegistry.default().convert_with_native(
        SourceBundle(root=tmp_path, primary=safe, companions=(companion,)),
        _svis(safe, "place"),
    )
    assert converted.native is not None
    assert any(record.collection == "external_choices" for record in converted.native.records)


def _known_item(name: str = "q") -> NativeRoutingItem:
    return NativeRoutingItem(
        local_id=f"item:{name}",
        source_item_id=name,
        raw_reference=name,
        label=name,
        section_path=(),
        source_order=0,
        block_ids=("block",),
        kind=NodeKind.question,
        parent_local_id=None,
        repeat_group_local_id=None,
        is_entry=False,
        linked_variable_names=(name,),
        source_text=name,
        terminal_kind=None,
        repeat_kind=None,
    )


@pytest.mark.parametrize(
    ("expression", "operator", "expected"),
    (
        ("${q} = 1", ConditionOperator.equals, 1),
        ("${q} != 2", ConditionOperator.not_equals, 2),
        ("${q} > 1.5", ConditionOperator.greater_than, 1.5),
        ("${q} >= true()", ConditionOperator.greater_than_or_equal, True),
        ("${q} < false()", ConditionOperator.less_than, False),
        ("${q} <= '2'", ConditionOperator.less_than_or_equal, "2"),
        ("${q} = ''", ConditionOperator.not_answered, None),
        ("${q} != ''", ConditionOperator.answered, None),
        ("selected(${q}, 'yes')", ConditionOperator.selected, "yes"),
        ("not(selected(${q}, 'yes'))", ConditionOperator.not_selected, "yes"),
        ("not (selected(${q}, 'yes'))", ConditionOperator.not_selected, "yes"),
        ("not(${q} = 1)", ConditionOperator.not_, None),
        ("(${q} = 1 or ${q} = 2)", ConditionOperator.any, None),
    ),
)
def test_xlsform_projects_the_exact_supported_operator_matrix(
    expression: str,
    operator: ConditionOperator,
    expected: object,
) -> None:
    projected = _project_expression(expression, {"q": _known_item()})
    assert projected.operator is operator
    if expected is not None:
        assert projected.value == expected


@pytest.mark.parametrize(
    "expression",
    (
        "${unknown} = 1",
        "selected(${q}, ${other})",
        "not(unsupported(${q}))",
        "${q} = 1 and unsupported(${q})",
        "${q} + 1 > 2",
    ),
)
def test_xlsform_keeps_unsupported_or_unresolved_expressions_opaque(expression: str) -> None:
    parsed = _parse_expression(expression, {"q": _known_item()})
    assert parsed.projection.operator is ConditionOperator.opaque
    assert parsed.exact_expression == expression


def test_expression_token_helpers_are_bounded_for_quotes_parentheses_and_bad_syntax() -> None:
    assert _split_top_level("${q} = 'a and b' and ${q} = 2", "and") == (
        "${q} = 'a and b'",
        "${q} = 2",
    )
    assert _split_top_level(") and ${q} = 2", "and") == (") and ${q} = 2",)
    assert _split_top_level("${q} = 1 and ", "and") == ("${q} = 1 and ",)
    assert _strip_outer_parentheses("(((${q} = 1)))") == "${q} = 1"
    assert _strip_outer_parentheses("(${q} = 1) trailing)") == "(${q} = 1) trailing)"
    assert _scalar("unquoted") is not True


def test_xlsform_table_and_type_helpers_reject_ambiguous_shapes() -> None:
    provenance = SourceProvenance(
        source_name="form.xlsx",
        sheet="survey",
        row_start=1,
        row_end=1,
    )
    with pytest.raises(SourceFormatError, match="header"):
        _table_records(SourceTable(id="empty", rows=(), provenance=provenance), required=("type",))
    duplicate = SourceTable(id="duplicate", rows=(("type", " TYPE "),), provenance=provenance)
    with pytest.raises(SourceFormatError, match="unique"):
        _table_records(duplicate, required=("type",))
    missing = SourceTable(id="missing", rows=(("name",),), provenance=provenance)
    with pytest.raises(SourceFormatError, match="required"):
        _table_records(missing, required=("type",))
    short = SourceTable(
        id="short",
        rows=(("type", "name", ""), ("text",), ("", "")),
        provenance=SourceProvenance(
            source_name="form.xlsx",
            sheet="survey",
            row_start=1,
            row_end=3,
        ),
    )
    headers, rows = _table_records(short, required=("type", "name"))
    assert headers == ("type", "name", "")
    assert rows[0][0] == 2
    assert rows[0][2] == {"type": "text", "name": ""}
    assert _choice_reference("text") is None
    assert _choice_reference("select_multiple list") == ("internal", "list")
    assert _choice_reference("select_multiple_from_file list.csv") == (
        "external",
        "list.csv",
    )
    assert _supported_question_type("datetime") is True
    assert _supported_question_type("unknown") is False


def test_non_xlsform_xlsx_keeps_document_conversion_and_no_native_payload(tmp_path: Path) -> None:
    path = tmp_path / "ordinary.xlsx"
    _survey_only(path, [["value"], ["one"]], title="Sheet1")
    converted = SourceRegistry.default().convert_with_native(path, _svis(path))
    assert converted.native is None
    assert converted.document.tables[0].rows == (("value",), ("one",))


@pytest.mark.parametrize(
    ("rows", "message"),
    (
        ([["type", "name"], ["end group", ""]], "unbalanced"),
        ([["type", "name"], ["text", ""]], "require names"),
        ([["type", "name"], ["text", "q"], ["integer", "q"]], "unique"),
        ([["type", "name"], ["begin group", "g"], ["text", "q"]], "unbalanced"),
        ([["type", "name"], ["begin repeat", "r"], ["end repeat", ""]], "at least one"),
    ),
)
def test_xlsform_rejects_malformed_group_and_item_structures(
    tmp_path: Path,
    rows: list[list[object]],
    message: str,
) -> None:
    path = tmp_path / "malformed.xlsx"
    _survey_only(path, rows)
    with pytest.raises(SourceFormatError, match=message):
        SourceRegistry.default().convert_with_native(path, _svis(path))


def test_xlsform_diagnoses_missing_choices_unknown_types_and_bind_relevance(
    tmp_path: Path,
) -> None:
    path = tmp_path / "diagnostics.xlsx"
    _survey_only(
        path,
        [
            ["type", "name", "label::French", "bind::relevant"],
            ["select_one missing", "q", "Question", ""],
            ["vendor_type", "other", "Autre", "${q} = 'yes'"],
        ],
    )
    converted = SourceRegistry.default().convert_with_native(path, _svis(path, "q", "other"))
    assert converted.native is not None
    assert [item.code for item in converted.native.diagnostics] == [
        "METADATA_INCOMPLETE",
        "XLSFORM_CHOICE_LIST_MISSING",
        "XLSFORM_TYPE_UNSUPPORTED",
    ]
    assert converted.native.items[1].label == "Question"
    assert (
        converted.native.activations[0].expression.projection.operator is ConditionOperator.equals
    )
    native_svis = SourceRegistry.default().convert_for_svis(
        path,
        extraction_date=date(2026, 9, 3),
    )
    assert native_svis.svis is not None
    assert native_svis.svis.variables[0].categories is None
    assert native_svis.svis.variables[0].needs_review is True


@pytest.mark.parametrize("filename", ("/outside.csv", "https://host/choices.csv", ""))
def test_external_choice_path_controls_cover_absolute_remote_and_empty_paths(
    tmp_path: Path,
    filename: str,
) -> None:
    path = tmp_path / "external.xlsx"
    _survey_only(
        path,
        [["type", "name"], [f"select_one_from_file {filename}", "q"]],
    )
    expected = SourceSecurityError if filename else SourceFormatError
    with pytest.raises(expected):
        SourceRegistry.default().convert_with_native(path, _svis(path, "q"))


def test_external_choice_missing_empty_and_aggregate_cell_limit(tmp_path: Path) -> None:
    path = tmp_path / "external.xlsx"
    _survey_only(path, [["type", "name"], ["select_one_from_file choices.csv", "q"]])
    with pytest.raises(SourceFormatError, match="missing"):
        SourceRegistry.default().convert_with_native(path, _svis(path, "q"))

    empty = tmp_path / "choices.csv"
    empty.write_text("", encoding="utf-8")
    converted = SourceRegistry.default().convert_with_native(
        SourceBundle(root=tmp_path, primary=path, companions=(empty,)),
        _svis(path, "q"),
    )
    assert converted.native is not None
    assert not any(record.collection == "external_choices" for record in converted.native.records)

    choices = tmp_path / "choices-full.csv"
    choices.write_text("name\na\nb\nc\n", encoding="utf-8")
    limited = tmp_path / "limited.xlsx"
    _survey_only(
        limited,
        [["type", "name"], ["select_one_from_file choices-full.csv", "q"]],
    )
    with pytest.raises(SourceLimitError, match="cell limit"):
        SourceRegistry.default().convert_with_native(
            SourceBundle(root=tmp_path, primary=limited, companions=(choices,)),
            _svis(limited, "q"),
            limits=replace(DEFAULT_SOURCE_LIMITS, max_cells=4),
        )


def test_xlsform_rejects_duplicate_internal_choice_identities(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-internal.xlsx"
    _workbook(path, [["type", "name"], ["select_one yes_no", "q"]])
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    workbook["choices"].append(["yes_no", "yes", "Duplicate", "Double"])
    workbook.save(path)
    workbook.close()

    with pytest.raises(SourceFormatError, match="choice identities must be unique"):
        SourceRegistry.default().convert_for_svis(path, extraction_date=date(2026, 9, 3))


def test_xlsform_rejects_duplicate_external_choice_identities(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-external.xlsx"
    companion = tmp_path / "choices.csv"
    _survey_only(path, [["type", "name"], ["select_one_from_file choices.csv", "q"]])
    companion.write_text("name,label\na,First\na,Duplicate\n", encoding="utf-8")

    with pytest.raises(SourceFormatError, match="external choice identities must be unique"):
        SourceRegistry.default().convert_for_svis(
            SourceBundle(root=tmp_path, primary=path, companions=(companion,)),
            extraction_date=date(2026, 9, 3),
        )


def test_xlsform_rejects_oversized_workbook_and_multiple_settings_rows(
    tmp_path: Path,
) -> None:
    oversized = tmp_path / "oversized.xlsx"
    _survey_only(oversized, [["type", "name"], ["text", "q"]])
    with pytest.raises(SourceLimitError, match="cell limit"):
        SourceRegistry.default().convert_with_native(
            oversized,
            _svis(oversized, "q"),
            limits=replace(DEFAULT_SOURCE_LIMITS, max_cells=3),
        )

    settings = tmp_path / "settings.xlsx"
    _workbook(settings, [["type", "name"], ["text", "q"]])
    from openpyxl import load_workbook

    workbook = load_workbook(settings)
    workbook["settings"].append(["Other", "other", "20260902"])
    workbook.save(settings)
    workbook.close()
    with pytest.raises(SourceFormatError, match="at most one"):
        SourceRegistry.default().convert_with_native(settings, _svis(settings, "q"))


def test_header_only_xlsform_routes_entry_directly_to_terminal(tmp_path: Path) -> None:
    path = tmp_path / "empty-form.xlsx"
    _survey_only(path, [["type", "name"]])
    converted = SourceRegistry.default().convert_with_native(path, _svis(path))
    assert converted.native is not None
    assert [item.kind for item in converted.native.items] == [NodeKind.entry, NodeKind.terminal]
    assert converted.native.transitions[0].target_local_id == "xlsform:terminal"


def test_versioned_fixture_builds_authoritative_typed_svis_without_provider_calls(
    tmp_path: Path,
) -> None:
    path, companion = _versioned_fixture(tmp_path)
    provider = DeterministicFakeProvider(
        capabilities=_capabilities(),
        responder=lambda _request: {"malicious": "replacement"},
    )

    result = SurveyScribe(provider, extraction_date=date(2026, 9, 3)).convert(
        SourceBundle(
            root=tmp_path,
            primary=Path(path.name),
            companions=(Path(companion.name),),
        )
    )

    assert result.output is not None
    assert type(result.output) is SurveySVIS
    assert provider.call_count == 0
    assert result.output.survey_id == "TST_2026_XLSFORM"
    assert result.output.survey_name == "Synthetic roster"
    assert result.output.country_code == "TST"
    assert result.output.year == 2026
    assert result.output.language == "French (fr)"
    assert result.output.data_collection_mode == "CAPI"
    assert [variable.raw_name for variable in result.output.variables] == [
        "consent",
        "age",
        "adult",
        "birthplace",
    ]
    consent, age, adult, birthplace = result.output.variables
    assert consent.label == "Consentement ?"
    assert consent.data_type is DataType.categorical_single
    assert [(category.code, category.label) for category in consent.categories or ()] == [
        ("yes", "Oui"),
        ("no", "Non"),
    ]
    assert age.module == "member"
    assert age.skip_condition_raw == "${consent} = 'yes'"
    assert age.notes == "XLSForm constraint: . >= 0"
    assert adult.notes == "XLSForm calculation: if(${age} >= 18, 1, 0)"
    assert [(category.code, category.label) for category in birthplace.categories or ()] == [
        ("capital", "Capitale"),
        ("rural", "Zone rurale"),
    ]

    bundle = SourceBundle(
        root=tmp_path,
        primary=Path(path.name),
        companions=(Path(companion.name),),
    )
    first = SourceRegistry.default().convert_for_svis(
        bundle,
        extraction_date=date(2026, 9, 3),
    )
    second = SourceRegistry.default().convert_for_svis(
        bundle,
        extraction_date=date(2026, 9, 3),
    )
    assert first == second
    assert first.native is not None
    assert all(item.raw_reference != "adult" for item in first.native.items)
    settings = next(record for record in first.native.records if record.collection == "settings")
    assert dict(settings.values)["version"] == "20260901"


@pytest.mark.parametrize(
    "rows",
    (
        [["type", "name"], ["text", "q"]],
        [["type", "name"]],
    ),
    ids=("missing-settings", "header-only"),
)
def test_public_xlsform_metadata_fallbacks_are_partial(
    tmp_path: Path,
    rows: list[list[object]],
) -> None:
    path = tmp_path / "fallback-form.xlsx"
    _survey_only(path, rows)
    provider = DeterministicFakeProvider(
        capabilities=_capabilities(),
        responder=lambda _request: {"malicious": "replacement"},
    )

    result = SurveyScribe(provider, extraction_date=date(2026, 9, 3)).convert(path)

    assert result.output is not None
    assert result.status is ResultStatus.partial
    assert provider.call_count == 0
    assert result.output.survey_id == "fallback-form"
    assert result.output.survey_name == "Fallback Form"
    assert result.output.country_code == "UNK"
    assert result.output.year == 2026
    assert [diagnostic.code for diagnostic in result.diagnostics] == [
        DiagnosticCode.metadata_incomplete
    ]


@pytest.mark.parametrize(
    "rows",
    (
        [["type", "name"], ["begin group", "g"], ["end repeat", ""]],
        [["type", "name"], ["begin repeat", "r"], ["end group", ""]],
    ),
)
def test_xlsform_rejects_mismatched_group_and_repeat_end_markers(
    tmp_path: Path,
    rows: list[list[object]],
) -> None:
    path = tmp_path / "mismatched.xlsx"
    _survey_only(path, rows)
    with pytest.raises(SourceFormatError, match="does not match"):
        SourceRegistry.default().convert_with_native(path, _svis(path))


def test_xlsform_emits_explicit_unsupported_and_unresolved_diagnostics(
    tmp_path: Path,
) -> None:
    path = tmp_path / "unsupported.xlsx"
    _survey_only(
        path,
        [
            ["type", "name", "relevant", "constraint", "appearance", "trigger"],
            ["integer", "age", "", "", "", ""],
            ["rank choices", "ranked", "count-selected(${age}) + 1 > 2", "", "", ""],
            ["text", "other", "${missing} = 1", "", "search('items')", "yes"],
        ],
    )

    converted = SourceRegistry.default().convert_with_native(
        path,
        _svis(path, "age", "ranked", "other"),
    )

    assert converted.native is not None
    assert {diagnostic.code for diagnostic in converted.native.diagnostics} == {
        "METADATA_INCOMPLETE",
        "XLSFORM_EXPRESSION_UNSUPPORTED",
        "XLSFORM_FEATURE_UNSUPPORTED",
        "XLSFORM_FUNCTION_UNSUPPORTED",
        "XLSFORM_REFERENCE_UNRESOLVED",
        "XLSFORM_TYPE_UNSUPPORTED",
    }
    native_svis = SourceRegistry.default().convert_for_svis(
        path,
        extraction_date=date(2026, 9, 3),
    )
    assert native_svis.svis is not None
    assert [variable.raw_name for variable in native_svis.svis.variables] == [
        "age",
        "ranked",
        "other",
    ]
    assert native_svis.svis.variables[1].data_type is DataType.other
    assert native_svis.svis.variables[1].needs_review is True

    public = SurveyScribe(
        DeterministicFakeProvider(
            capabilities=_capabilities(),
            responder=lambda _request: {"malicious": "replacement"},
        ),
        extraction_date=date(2026, 9, 3),
    ).convert(path)
    assert {diagnostic.code for diagnostic in public.diagnostics} == {
        DiagnosticCode.metadata_incomplete,
        DiagnosticCode.xlsform_expression_unsupported,
        DiagnosticCode.xlsform_feature_unsupported,
        DiagnosticCode.xlsform_function_unsupported,
        DiagnosticCode.xlsform_reference_unresolved,
        DiagnosticCode.xlsform_type_unsupported,
    }


def test_xlsform_rejects_corrupt_package(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not an xlsx package")
    with pytest.raises(SourceFormatError, match="extension does not match"):
        SourceRegistry.default().convert_with_native(path, _svis(path))


def test_xlsform_uses_one_deadline_for_workbook_and_external_csv(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path, companion = _versioned_fixture(tmp_path)
    adapter = XlsFormAdapter()
    resolved = SourceRegistry.default().convert(
        path,
        limits=replace(DEFAULT_SOURCE_LIMITS, deadline_seconds=10),
    )
    captured: list[float] = []

    def expire_csv(
        self: object,
        source: object,
        *,
        limits: object,
        deadline: float,
    ) -> object:
        del self, source, limits
        captured.append(deadline)
        raise SourceTimeoutError("Source conversion exceeded the configured deadline")

    monkeypatch.setattr("survey_scribe.sources.xlsform.CsvAdapter.convert_until", expire_csv)
    deadline = time.monotonic() + 10
    from survey_scribe.sources.base import resolve_local_source

    source = resolve_local_source(
        SourceBundle(
            root=tmp_path,
            primary=Path(path.name),
            companions=(Path(companion.name),),
        )
    )
    with pytest.raises(SourceTimeoutError):
        adapter.convert_svis_until(
            source,
            resolved,
            limits=DEFAULT_SOURCE_LIMITS,
            deadline=deadline,
            extraction_date=date(2026, 9, 3),
        )
    assert captured == [deadline]

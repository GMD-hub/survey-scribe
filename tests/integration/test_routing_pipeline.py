"""Public routed pipeline integration, native bypass, and outcome tests."""

from __future__ import annotations

import asyncio
import json
from datetime import date
from pathlib import Path
from threading import Event as ThreadEvent
from typing import cast

import pytest

from survey_scribe.models.routing import RoutingSourceBinding
from survey_scribe.models.svis import AnswerCategory, DataType, SurveySVIS, SurveyVariable
from survey_scribe.providers.base import ProviderTransportError
from survey_scribe.providers.capabilities import CapabilityEvidence, ModelCapabilities
from survey_scribe.providers.testing import DeterministicFakeProvider, FakeRequest
from survey_scribe.results import ResultStatus
from survey_scribe.routing import pipeline as pipeline_module
from survey_scribe.routing.contracts import ConditionOperator, RoutingEvidenceBatch, RoutingPassKind
from survey_scribe.routing.pipeline import QuestionnaireRouter
from survey_scribe.serialization.routing import ArtifactManifestV2, parse_artifact_manifest
from survey_scribe.sources.base import SourceBundle
from survey_scribe.sources.registry import SourceConversionResult, SourceRegistry


def _workbook(path: Path, *, opaque: bool = False) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    survey = workbook.active
    assert survey is not None
    survey.title = "survey"
    survey.append(["type", "name", "label", "relevant"])
    survey.append(["select_one yes_no", "consent", "Consent?", ""])
    survey.append(["begin repeat", "member", "Member", "${consent} = 'yes'"])
    survey.append(
        [
            "integer",
            "age",
            "Age",
            "count-selected(${consent}) + 1 > 2" if opaque else "${consent} = 'yes'",
        ]
    )
    survey.append(["end repeat", "", "", ""])
    choices = workbook.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    choices.append(["yes_no", "yes", "Yes"])
    choices.append(["yes_no", "no", "No"])
    settings = workbook.create_sheet("settings")
    settings.append(["form_title", "form_id"])
    settings.append(["Roster", "roster"])
    workbook.save(path)


def _svis(path: Path, *raw_names: str) -> SurveySVIS:
    return SurveySVIS(
        survey_id="TST_2026_ROUTE",
        country_code="TST",
        year=2026,
        survey_name="Synthetic routed survey",
        variables=[
            SurveyVariable(
                raw_name=name,
                label=name.title(),
                data_type=DataType.numeric,
                extraction_confidence=1.0,
            )
            for name in raw_names
        ],
        source_file=path.name,
        source_format=path.suffix.removeprefix("."),
        extraction_date=date(2026, 9, 1),
    )


def _binding(path: Path, svis: SurveySVIS) -> RoutingSourceBinding:
    return SourceRegistry.default().convert_with_native(path, svis).source_binding


def _capabilities() -> ModelCapabilities:
    return ModelCapabilities(
        provider="fake",
        model="routing-fake-v1",
        structured_output=True,
        strict_schema=True,
        max_input_tokens=200_000,
        max_output_tokens=4_096,
        supported_generation_settings=frozenset({"temperature", "max_output_tokens", "seed"}),
        evidence=CapabilityEvidence.verified,
        tested_sdk_version="fake-1",
    )


def _data_block(content: str, name: str) -> object:
    start = f"BEGIN_UNTRUSTED_{name}_JSON\n"
    end = f"\nEND_UNTRUSTED_{name}_JSON"
    return json.loads(content.split(start, maxsplit=1)[1].split(end, maxsplit=1)[0])


async def _empty_responder(request: FakeRequest) -> RoutingEvidenceBatch:
    task = next(message.content for message in request.messages if message.role == "user")
    chunk_id = cast(str, json.loads(task.split("CHUNK_JSON: ", maxsplit=1)[1].splitlines()[0]))
    if task.startswith("PASS: forward"):
        inventory = cast(list[dict[str, object]], _data_block(task, "ITEM_INVENTORY"))
        pass_kind = RoutingPassKind.forward
    else:
        inventory = cast(list[dict[str, object]], _data_block(task, "TARGET_ITEMS"))
        pass_kind = RoutingPassKind.incoming_activation
    return RoutingEvidenceBatch(
        chunk_id=chunk_id,
        pass_kind=pass_kind,
        examined_item_ids=tuple(
            cast(str, item["source_item_id"] or item["raw_reference"]) for item in inventory
        ),
        evidence=(),
        unresolved_references=(),
        notes=(),
    )


def _provider() -> DeterministicFakeProvider:
    return DeterministicFakeProvider(capabilities=_capabilities(), responder=_empty_responder)


class _StaticRegistry:
    def __init__(self, result: SourceConversionResult) -> None:
        self.result = result

    def convert_with_native(self, *_args: object, **_kwargs: object) -> SourceConversionResult:
        return self.result


def _static_registry(result: SourceConversionResult) -> SourceRegistry:
    return cast(SourceRegistry, _StaticRegistry(result))


@pytest.mark.asyncio
async def test_real_native_xlsform_routes_relevance_and_repeat_with_zero_model_calls(
    tmp_path: Path,
) -> None:
    path = tmp_path / "roster.xlsx"
    _workbook(path)
    svis = _svis(path, "consent", "age")
    provider = _provider()

    result = await QuestionnaireRouter(provider).aroute(
        path,
        svis,
        source_binding=_binding(path, svis),
    )

    assert result.status is ResultStatus.success
    assert result.output is not None
    assert provider.call_count == 0
    assert [variable.raw_name for variable in result.output.variables] == ["consent", "age"]
    assert all(variable.routing_node_id is not None for variable in result.output.variables)
    assert len(result.output.routing_graph.loops) == 1
    repeat = next(
        node for node in result.output.routing_graph.nodes if node.kind.value == "repeat_group"
    )
    assert repeat.repeat_spec is not None
    assert repeat.containment.entry_child_node_id is not None
    age = next(node for node in result.output.routing_graph.nodes if node.raw_name == "age")
    assert age.activation_condition is not None
    assert age.activation_condition.operator is ConditionOperator.equals
    native_evidence = result.output.routing_graph.routing_audit.evidence
    assert native_evidence
    assert all(record.observation.native_expression is not None for record in native_evidence)


def test_native_no_provider_sync_async_parity_and_running_loop_rejection(tmp_path: Path) -> None:
    path = tmp_path / "roster.xlsx"
    _workbook(path)
    svis = _svis(path, "consent", "age")
    binding = _binding(path, svis)
    router = QuestionnaireRouter(None)

    sync_result = router.route(path, svis, source_binding=binding)
    async_result = asyncio.run(router.aroute(path, svis, source_binding=binding))

    assert sync_result.output == async_result.output
    assert sync_result.status is ResultStatus.success

    async def reject() -> None:
        with pytest.raises(RuntimeError, match="running event loop"):
            router.route(path, svis, source_binding=binding)

    asyncio.run(reject())


@pytest.mark.asyncio
async def test_complete_binding_mismatch_fails_before_provider_or_inventory_work(
    tmp_path: Path,
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age\nQ2. Done", encoding="utf-8")
    svis = _svis(path, "age")
    provider = _provider()
    correct = _binding(path, svis)
    variants = (
        correct.model_copy(update={"survey_id": "OTHER"}),
        correct.model_copy(update={"source_name": "other.txt"}),
        correct.model_copy(update={"media_type": "application/pdf"}),
        correct.model_copy(update={"snapshot_sha256": "f" * 64}),
        correct.model_copy(update={"source_conversion_schema_version": "2.0"}),
    )

    for binding in variants:
        result = await QuestionnaireRouter(provider).aroute(
            path,
            svis,
            source_binding=binding,
        )
        assert result.status is ResultStatus.failed
        assert result.output is None
        assert [diagnostic.code for diagnostic in result.diagnostics] == ["ROUTING_SOURCE_MISMATCH"]
    assert provider.call_count == 0


@pytest.mark.asyncio
async def test_companion_mutation_fails_binding_before_provider_call(tmp_path: Path) -> None:
    path = tmp_path / "questionnaire.txt"
    companion = tmp_path / "labels.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    companion.write_text("Age label", encoding="utf-8")
    bundle = SourceBundle(
        root=tmp_path,
        primary=Path(path.name),
        companions=(Path(companion.name),),
    )
    svis = _svis(path, "age")
    binding = SourceRegistry.default().convert_with_native(bundle, svis).source_binding
    companion.write_text("Changed age label", encoding="utf-8")
    provider = _provider()

    result = await QuestionnaireRouter(provider).aroute(
        bundle,
        svis,
        source_binding=binding,
    )

    assert result.status is ResultStatus.failed
    assert result.diagnostics[0].code == "ROUTING_SOURCE_MISMATCH"
    assert provider.call_count == 0


@pytest.mark.asyncio
async def test_public_router_passes_linked_svis_category_codes_to_validation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "age")
    svis.variables[0].categories = [
        AnswerCategory(code=1, label="One"),
        AnswerCategory(code=2, label="Two"),
    ]
    captured: dict[str, tuple[object, ...]] = {}
    original_extract = pipeline_module.extract_routing

    async def capture_categories(**kwargs: object) -> object:
        captured.update(kwargs["known_category_codes"])  # type: ignore[arg-type]
        return await original_extract(**kwargs)  # type: ignore[arg-type]

    monkeypatch.setattr(pipeline_module, "extract_routing", capture_categories)

    result = await QuestionnaireRouter(_provider()).aroute(
        path,
        svis,
        source_binding=_binding(path, svis),
    )

    assert result.output is not None
    assert tuple(captured.values()) == ((1, 2),)


@pytest.mark.asyncio
async def test_non_native_source_requires_provider_and_provider_path_is_stable(
    tmp_path: Path,
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age\nQ2. Employment", encoding="utf-8")
    svis = _svis(path, "age", "employment")
    binding = _binding(path, svis)

    absent = await QuestionnaireRouter(None).aroute(path, svis, source_binding=binding)
    assert absent.status is ResultStatus.failed
    assert absent.diagnostics[0].code == "ROUTING_PROVIDER_REQUIRED"

    provider = _provider()
    routed = await QuestionnaireRouter(provider).aroute(path, svis, source_binding=binding)
    assert routed.status is ResultStatus.success
    assert routed.output is not None
    assert provider.call_count > 0
    assert [item.raw_name for item in routed.output.variables] == ["age", "employment"]
    linked_nodes = {
        node.node_id: node
        for node in routed.output.routing_graph.nodes
        if node.kind.value == "question"
    }
    assert [
        linked_nodes[item.routing_node_id].source_item_id
        for item in routed.output.variables
        if item.routing_node_id is not None
    ] == ["Q1", "Q2"]
    orders = [item.source_order for item in routed.output.routing_graph.routing_audit.inventory]
    assert orders == sorted(orders)
    assert routed.artifact_provenance is not None
    assert routed.artifact_provenance.source_sha256 == (binding.snapshot_sha256,)
    assert routed.artifact_provenance.model_response_sha256
    assert routed.artifact_provenance.prompt_versions

    written = routed.write(tmp_path)
    manifest_path = next(item.path for item in written.artifacts if item.kind == "manifest")
    manifest = parse_artifact_manifest(manifest_path.read_bytes())
    assert isinstance(manifest, ArtifactManifestV2)
    assert manifest.source_sha256 == routed.artifact_provenance.source_sha256
    assert manifest.model_response_sha256 == routed.artifact_provenance.model_response_sha256
    assert manifest.prompt_versions == routed.artifact_provenance.prompt_versions


@pytest.mark.asyncio
async def test_document_variables_link_only_to_unique_exact_source_identity(tmp_path: Path) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q9. Routing instruction\nQ1. Age", encoding="utf-8")
    svis = _svis(path, "age", "employment")

    result = await QuestionnaireRouter(_provider()).aroute(
        path,
        svis,
        source_binding=_binding(path, svis),
    )

    assert result.status is ResultStatus.partial
    assert result.output is not None
    age, employment = result.output.variables
    assert age.routing_node_id is not None
    assert employment.routing_node_id is None
    nodes = {node.node_id: node for node in result.output.routing_graph.nodes}
    assert nodes[age.routing_node_id].source_item_id == "Q1"
    q9 = next(
        item
        for item in result.output.routing_graph.routing_audit.inventory
        if item.source_item_id == "Q9"
    )
    assert q9.linked_variable_indices == ()
    assert any(diagnostic.code == "UNLINKED_VARIABLE" for diagnostic in result.diagnostics)


@pytest.mark.asyncio
async def test_printed_identity_precedes_competing_variable_label(tmp_path: Path) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "Q1", "age")

    result = await QuestionnaireRouter(_provider()).aroute(
        path,
        svis,
        source_binding=_binding(path, svis),
    )

    assert result.status is ResultStatus.partial
    assert result.output is not None
    printed, label_only = result.output.variables
    assert printed.routing_node_id is not None
    assert label_only.routing_node_id is None
    node = next(
        item
        for item in result.output.routing_graph.nodes
        if item.node_id == printed.routing_node_id
    )
    assert node.source_item_id == "Q1"


@pytest.mark.asyncio
async def test_incomplete_native_source_uses_provider_enrichment_and_partial_failure(
    tmp_path: Path,
) -> None:
    path = tmp_path / "mixed.xlsx"
    _workbook(path)
    svis = _svis(path, "consent", "age")
    conversion = SourceRegistry.default().convert_with_native(path, svis)
    assert conversion.native is not None
    incomplete = SourceConversionResult(
        document=conversion.document,
        source_binding=conversion.source_binding,
        native=conversion.native.model_copy(update={"complete": False}),
    )

    absent = await QuestionnaireRouter(
        None,
        sources=_static_registry(incomplete),
    ).aroute(path, svis, source_binding=conversion.source_binding)
    assert absent.status is ResultStatus.failed
    assert absent.diagnostics[0].code == "ROUTING_PROVIDER_REQUIRED"

    provider = _provider()
    mixed = await QuestionnaireRouter(
        provider,
        sources=_static_registry(incomplete),
    ).aroute(path, svis, source_binding=conversion.source_binding)
    assert mixed.status is ResultStatus.success
    assert mixed.output is not None
    assert provider.call_count > 0
    assert mixed.output.routing_graph.routing_audit.evidence

    async def fail(_request: FakeRequest) -> object:
        raise ProviderTransportError(retryable=False)

    failed_provider = DeterministicFakeProvider(capabilities=_capabilities(), responder=fail)
    partial = await QuestionnaireRouter(
        failed_provider,
        sources=_static_registry(incomplete),
    ).aroute(path, svis, source_binding=conversion.source_binding)
    assert partial.status is ResultStatus.partial
    assert partial.output is not None
    assert partial.failed_blocks[0].block_id == "forward-000001"


@pytest.mark.asyncio
async def test_empty_inventory_source_error_and_total_provider_failure_are_failed(
    tmp_path: Path,
) -> None:
    empty_path = tmp_path / "empty.txt"
    empty_path.write_text("", encoding="utf-8")
    empty_svis = _svis(empty_path)
    empty_binding = _binding(empty_path, empty_svis)
    empty = await QuestionnaireRouter(_provider()).aroute(
        empty_path,
        empty_svis,
        source_binding=empty_binding,
    )
    assert empty.status is ResultStatus.failed
    assert empty.diagnostics[0].code == "ROUTING_EMPTY_INVENTORY"

    unsupported = tmp_path / "questionnaire.unsupported"
    unsupported.write_text("Q1", encoding="utf-8")
    placeholder = RoutingSourceBinding(
        survey_id="TST_2026_ROUTE",
        source_name=unsupported.name,
        media_type="application/octet-stream",
        snapshot_sha256="a" * 64,
        source_conversion_schema_version="1.0",
    )
    source_error = await QuestionnaireRouter(_provider()).aroute(
        unsupported,
        _svis(unsupported, "q"),
        source_binding=placeholder,
    )
    assert source_error.status is ResultStatus.failed
    assert source_error.diagnostics[0].code == "SOURCE_FORMAT_UNSUPPORTED"

    text_path = tmp_path / "provider.txt"
    text_path.write_text("Q1. Age", encoding="utf-8")
    text_svis = _svis(text_path, "age")

    async def fail(_request: FakeRequest) -> object:
        raise ProviderTransportError(retryable=False)

    failed_provider = DeterministicFakeProvider(capabilities=_capabilities(), responder=fail)
    provider_failure = await QuestionnaireRouter(failed_provider).aroute(
        text_path,
        text_svis,
        source_binding=_binding(text_path, text_svis),
    )
    assert provider_failure.status is ResultStatus.failed
    assert provider_failure.output is None
    assert provider_failure.diagnostics[0].code == "ROUTING_PROVIDER_TRANSPORT"


@pytest.mark.asyncio
async def test_document_inventory_fallbacks_and_exception_boundaries_are_safe(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "plain.txt"
    path.write_text("Age question without a printed identifier", encoding="utf-8")
    svis = _svis(path, "age", "employment")
    svis.variables[0].question_text = "Age question"
    svis.variables[1].label = "Employment"
    binding = _binding(path, svis)
    successful = await QuestionnaireRouter(_provider()).aroute(
        path,
        svis,
        source_binding=binding,
    )
    assert successful.output is not None
    inventory = successful.output.routing_graph.routing_audit.inventory
    assert all(item.source_item_id is None for item in inventory)

    async def fail_before_graph(**_kwargs: object) -> object:
        raise ValueError("private provider detail")

    monkeypatch.setattr("survey_scribe.routing.pipeline.extract_routing", fail_before_graph)
    with pytest.raises(ValueError, match="private provider detail"):
        await QuestionnaireRouter(_provider()).aroute(path, svis, source_binding=binding)

    async def expected_provider_failure(**_kwargs: object) -> object:
        raise ProviderTransportError(retryable=False)

    monkeypatch.setattr(
        "survey_scribe.routing.pipeline.extract_routing",
        expected_provider_failure,
    )
    failed = await QuestionnaireRouter(_provider()).aroute(path, svis, source_binding=binding)
    assert failed.status is ResultStatus.failed
    assert failed.diagnostics[0].code == "ROUTING_PROVIDER_FAILED"
    assert "private provider detail" not in repr(failed)


@pytest.mark.asyncio
async def test_source_binding_validation_error_is_normalized_to_mismatch(tmp_path: Path) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "age")
    svis.source_file = "other.txt"
    placeholder = RoutingSourceBinding(
        survey_id=svis.survey_id,
        source_name=path.name,
        media_type="text/plain",
        snapshot_sha256="a" * 64,
        source_conversion_schema_version="1.0",
    )
    result = await QuestionnaireRouter(_provider()).aroute(
        path,
        svis,
        source_binding=placeholder,
    )
    assert result.status is ResultStatus.failed
    assert result.diagnostics[0].code == "ROUTING_SOURCE_MISMATCH"


@pytest.mark.asyncio
async def test_unlinked_variable_is_partial_and_opaque_native_warning_is_success(
    tmp_path: Path,
) -> None:
    path = tmp_path / "roster.xlsx"
    _workbook(path)
    unlinked_svis = _svis(path, "consent", "age", "not_in_form")
    partial = await QuestionnaireRouter(None).aroute(
        path,
        unlinked_svis,
        source_binding=_binding(path, unlinked_svis),
    )
    assert partial.status is ResultStatus.partial
    assert partial.output is not None
    assert partial.output.variables[-1].routing_node_id is None
    assert any(diagnostic.code == "UNLINKED_VARIABLE" for diagnostic in partial.diagnostics)

    opaque_path = tmp_path / "opaque.xlsx"
    _workbook(opaque_path, opaque=True)
    opaque_svis = _svis(opaque_path, "consent", "age")
    warning = await QuestionnaireRouter(None).aroute(
        opaque_path,
        opaque_svis,
        source_binding=_binding(opaque_path, opaque_svis),
    )
    assert warning.status is ResultStatus.success
    assert warning.output is not None
    expression = next(
        record.observation.native_expression
        for record in warning.output.routing_graph.routing_audit.evidence
        if record.observation.native_expression is not None
        and record.observation.native_expression.exact_expression.startswith("count-selected")
    )
    assert expression.canonical_projection.operator is ConditionOperator.opaque
    assert any(
        diagnostic.code == "OPAQUE_ACTIVATION_CONDITION"
        for diagnostic in warning.output.routing_graph.diagnostics
    )


@pytest.mark.asyncio
async def test_cancellation_propagates_from_provider_work(
    tmp_path: Path,
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "age")
    control = asyncio.CancelledError()

    async def stop(_request: FakeRequest) -> object:
        raise control

    provider = DeterministicFakeProvider(capabilities=_capabilities(), responder=stop)
    with pytest.raises(type(control)):
        await QuestionnaireRouter(provider).aroute(
            path,
            svis,
            source_binding=_binding(path, svis),
        )


@pytest.mark.asyncio
async def test_source_conversion_runs_off_loop_and_cancellation_prevents_provider_call(
    tmp_path: Path,
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "age")
    conversion = SourceRegistry.default().convert_with_native(path, svis)
    entered = ThreadEvent()
    release = ThreadEvent()
    finished = ThreadEvent()

    class BlockingRegistry:
        def convert_with_native(self, *_args: object, **_kwargs: object) -> SourceConversionResult:
            entered.set()
            release.wait(timeout=5)
            finished.set()
            return conversion

    provider = _provider()
    router = QuestionnaireRouter(provider, sources=cast(SourceRegistry, BlockingRegistry()))
    task = asyncio.create_task(router.aroute(path, svis, source_binding=conversion.source_binding))
    assert await asyncio.to_thread(entered.wait, 2)
    assert finished.is_set() is False

    ticked = False

    async def tick() -> None:
        nonlocal ticked
        await asyncio.sleep(0)
        ticked = True

    await tick()
    assert ticked is True
    task.cancel()
    try:
        with pytest.raises(asyncio.CancelledError):
            await task
    finally:
        release.set()
    assert provider.call_count == 0


@pytest.mark.asyncio
@pytest.mark.parametrize("control_type", (KeyboardInterrupt, SystemExit))
async def test_process_control_propagates_from_orchestration(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    control_type: type[BaseException],
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "age")

    async def stop(**_kwargs: object) -> object:
        raise control_type

    monkeypatch.setattr("survey_scribe.routing.pipeline.extract_routing", stop)
    with pytest.raises(control_type):
        await QuestionnaireRouter(_provider()).aroute(
            path,
            svis,
            source_binding=_binding(path, svis),
        )


@pytest.mark.asyncio
async def test_svis_is_detached_before_provider_work_and_injected_objects_are_not_closed(
    tmp_path: Path,
) -> None:
    path = tmp_path / "questionnaire.txt"
    path.write_text("Q1. Age", encoding="utf-8")
    svis = _svis(path, "age")
    binding = _binding(path, svis)
    entered = asyncio.Event()
    release = asyncio.Event()

    async def wait_then_respond(request: FakeRequest) -> RoutingEvidenceBatch:
        entered.set()
        await release.wait()
        return await _empty_responder(request)

    provider = DeterministicFakeProvider(
        capabilities=_capabilities(),
        responder=wait_then_respond,
    )
    closed: list[str] = []
    provider.close = lambda: closed.append("provider")  # type: ignore[attr-defined]
    registry = SourceRegistry.default()
    registry.close = lambda: closed.append("registry")  # type: ignore[attr-defined]
    router = QuestionnaireRouter(provider, sources=registry)
    task = asyncio.create_task(router.aroute(path, svis, source_binding=binding))
    await entered.wait()
    original_name = svis.survey_name
    svis.survey_name = "Mutated after entry"
    svis.variables.reverse()
    release.set()

    result = await task

    assert result.output is not None
    assert result.output.survey_name == original_name
    assert [item.raw_name for item in result.output.variables] == ["age"]
    assert closed == []


def test_questionnaire_router_public_signatures_are_exact() -> None:
    import inspect

    constructor = inspect.signature(QuestionnaireRouter)
    route = inspect.signature(QuestionnaireRouter.route)
    aroute = inspect.signature(QuestionnaireRouter.aroute)

    assert tuple(constructor.parameters) == ("provider", "config", "sources")
    assert constructor.parameters["config"].kind is inspect.Parameter.KEYWORD_ONLY
    assert constructor.parameters["sources"].kind is inspect.Parameter.KEYWORD_ONLY
    assert tuple(route.parameters) == ("self", "source", "svis", "source_binding")
    assert route.parameters["source_binding"].kind is inspect.Parameter.KEYWORD_ONLY
    assert tuple(aroute.parameters) == ("self", "source", "svis", "source_binding")
    assert aroute.parameters["source_binding"].kind is inspect.Parameter.KEYWORD_ONLY
